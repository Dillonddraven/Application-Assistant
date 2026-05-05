"""Persistence for the finder review queue. Mirrors the tracker XLSX style
so the user can open it in Excel/Numbers and act on it.

V1 usage-upgrade schema adds grouping (`company_group_id`, `company_rank`,
`role_rank_within_company`), categorization (`role_category`), pre-tailor
estimates (`production_status_estimate`, `strategic_interest`), reason
strings (`why_matched`, `why_risky`), and the user-editable approval columns
(`approved_for_packet`, `outreach_group_id`, `company_outreach_needed`,
`user_notes`)."""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config import ROOT

QUEUE_PATH = ROOT / "runs" / "_finder_queue.xlsx"

COLUMNS = [
    # Identity
    "company", "title", "location", "ats", "url",
    # Scoring
    "quick_fit", "stretch_flags",
    # Categorization (added v1-usage)
    "role_category",                # core_security | security_adjacent | stretch | non_core
    # Recommendations
    "recommended_action",           # run_packet | save_for_later | skip | stretch
    "recommended_next_action",      # run_packet | bundle_with_company | save_for_later | skip
    "production_status_estimate",   # portal_ready | maybe_review | skip
    "strategic_interest",           # high | medium | low
    "why_matched", "why_risky",
    # Grouping (added v1-usage)
    "company_group_id",             # stable id (== ats_slug)
    "company_rank",                 # 1 = best company
    "role_rank_within_company",     # 1 = best role within company
    # Apply links
    "apply_url",                    # where to apply (defaults to url)
    "direct_apply_url",             # user-editable; if non-empty overrides apply_url
    "source_url",                   # original posting URL (== url)
    # User-editable workflow columns
    "review_status",                # new | reviewed | approved | ignored
    "approved_for_packet",          # y | n | "" (user marks for packet generation)
    "outreach_group_id",            # company_group_id once bundled
    "company_outreach_needed",      # y if company has 2+ matched roles
    "user_notes",
    # Ranking (added v1.1)
    "ranking_score",                # 0-100 weighted (35/25/20/10/10)
    "ranking_category",             # Strong Target | Target | Reach | Too Senior | Blocker
    "experience_level",             # entry | junior | target | mid | senior | unknown
    "salary_bucket",                # target | reach | mid_senior | senior | below_target | unknown
    "role_family_label",            # target | adjacent | security_other | off-target
    "blockers_summary",             # short string of blocker.kind:detail; empty if none
    "biggest_reason",
    "biggest_risk",
    "resume_angle",
    "outreach_angle",
    "tailor_resume",                # y | n
    "final_recommendation",         # Apply | Maybe apply | Skip
    # Misc
    "discovered_date", "review_notes", "external_id",
]

ACTION_COLORS = {
    "run_packet":             PatternFill("solid", fgColor="DDF0D2"),
    "bundle_with_company":    PatternFill("solid", fgColor="C7E6F5"),
    "save_for_later":         PatternFill("solid", fgColor="FFF7D6"),
    "stretch":                PatternFill("solid", fgColor="DDE8F7"),
    "skip":                   PatternFill("solid", fgColor="EAD6D6"),
}
CATEGORY_COLORS = {
    "core_security":     PatternFill("solid", fgColor="DDF0D2"),
    "security_adjacent": PatternFill("solid", fgColor="C7E6F5"),
    "stretch":           PatternFill("solid", fgColor="DDE8F7"),
    "non_core":          PatternFill("solid", fgColor="EAD6D6"),
}
INTEREST_COLORS = {
    "high":   PatternFill("solid", fgColor="DDF0D2"),
    "medium": PatternFill("solid", fgColor="FFF7D6"),
    "low":    PatternFill("solid", fgColor="EAD6D6"),
}
HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT = Font(bold=True, color="FFFFFF")


@dataclass
class FinderQueueRow:
    company: str = ""
    title: str = ""
    location: str = ""
    ats: str = ""
    url: str = ""
    quick_fit: int = 0
    stretch_flags: str = ""
    role_category: str = ""
    recommended_action: str = "save_for_later"
    recommended_next_action: str = ""
    production_status_estimate: str = ""
    strategic_interest: str = ""
    why_matched: str = ""
    why_risky: str = ""
    company_group_id: str = ""
    company_rank: int = 0
    role_rank_within_company: int = 0
    apply_url: str = ""
    direct_apply_url: str = ""
    source_url: str = ""
    review_status: str = "new"
    approved_for_packet: str = ""
    outreach_group_id: str = ""
    company_outreach_needed: str = ""
    user_notes: str = ""
    discovered_date: str = ""
    review_notes: str = ""
    external_id: str = ""
    # Ranking (v1.1)
    ranking_score: int = 0
    ranking_category: str = ""
    experience_level: str = ""
    salary_bucket: str = ""
    role_family_label: str = ""
    blockers_summary: str = ""
    biggest_reason: str = ""
    biggest_risk: str = ""
    resume_angle: str = ""
    outreach_angle: str = ""
    tailor_resume: str = ""           # "y" | "n"
    final_recommendation: str = ""

    def as_dict(self) -> dict[str, Any]:
        return {col: getattr(self, col) for col in COLUMNS}


# Fields the user may edit in the XLSX that the runner must NOT overwrite on
# re-runs. Everything else is recomputed from the freshly-fetched posting.
USER_EDITABLE_FIELDS = (
    "review_status", "review_notes", "approved_for_packet",
    "outreach_group_id", "user_notes", "direct_apply_url",
)


def _fresh_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "finder_queue"
    for i, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"
    widths = {
        "company": 22, "title": 36, "location": 18, "ats": 10,
        "url": 50, "quick_fit": 10, "stretch_flags": 28,
        "role_category": 18,
        "recommended_action": 18, "recommended_next_action": 22,
        "production_status_estimate": 22, "strategic_interest": 16,
        "why_matched": 38, "why_risky": 28,
        "company_group_id": 16, "company_rank": 12,
        "role_rank_within_company": 14,
        "apply_url": 50, "direct_apply_url": 40, "source_url": 50,
        "review_status": 14, "approved_for_packet": 18,
        "outreach_group_id": 18, "company_outreach_needed": 20,
        "user_notes": 30,
        "discovered_date": 12, "review_notes": 30, "external_id": 16,
    }
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 14)
    return wb


def _coerce(val, *, default=""):
    if val is None:
        return default
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    return val


def load_queue(path: Path | None = None) -> list[FinderQueueRow]:
    p = path or QUEUE_PATH
    if not p.exists():
        return []
    wb = openpyxl.load_workbook(str(p))
    ws = wb["finder_queue"] if "finder_queue" in wb.sheetnames else wb.active
    out: list[FinderQueueRow] = []
    # Build a header map so older queues with fewer columns still load.
    header_map: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        name = ws.cell(row=1, column=col_idx).value
        if isinstance(name, str):
            header_map[name] = col_idx
    for r in range(2, ws.max_row + 1):
        vals: dict[str, Any] = {}
        for col in COLUMNS:
            ci = header_map.get(col)
            v = _coerce(ws.cell(row=r, column=ci).value) if ci else ""
            vals[col] = v if v != "" else ""
        if not vals.get("url") and not vals.get("company"):
            continue
        out.append(FinderQueueRow(
            company=str(vals.get("company") or ""),
            title=str(vals.get("title") or ""),
            location=str(vals.get("location") or ""),
            ats=str(vals.get("ats") or ""),
            url=str(vals.get("url") or ""),
            quick_fit=int(vals.get("quick_fit") or 0),
            stretch_flags=str(vals.get("stretch_flags") or ""),
            role_category=str(vals.get("role_category") or ""),
            recommended_action=str(vals.get("recommended_action") or "save_for_later"),
            recommended_next_action=str(vals.get("recommended_next_action") or ""),
            production_status_estimate=str(vals.get("production_status_estimate") or ""),
            strategic_interest=str(vals.get("strategic_interest") or ""),
            why_matched=str(vals.get("why_matched") or ""),
            why_risky=str(vals.get("why_risky") or ""),
            company_group_id=str(vals.get("company_group_id") or ""),
            company_rank=int(vals.get("company_rank") or 0),
            role_rank_within_company=int(vals.get("role_rank_within_company") or 0),
            apply_url=str(vals.get("apply_url") or ""),
            direct_apply_url=str(vals.get("direct_apply_url") or ""),
            source_url=str(vals.get("source_url") or ""),
            review_status=str(vals.get("review_status") or "new"),
            approved_for_packet=str(vals.get("approved_for_packet") or ""),
            outreach_group_id=str(vals.get("outreach_group_id") or ""),
            company_outreach_needed=str(vals.get("company_outreach_needed") or ""),
            user_notes=str(vals.get("user_notes") or ""),
            discovered_date=str(vals.get("discovered_date") or ""),
            review_notes=str(vals.get("review_notes") or ""),
            external_id=str(vals.get("external_id") or ""),
            ranking_score=int(vals.get("ranking_score") or 0),
            ranking_category=str(vals.get("ranking_category") or ""),
            experience_level=str(vals.get("experience_level") or ""),
            salary_bucket=str(vals.get("salary_bucket") or ""),
            role_family_label=str(vals.get("role_family_label") or ""),
            blockers_summary=str(vals.get("blockers_summary") or ""),
            biggest_reason=str(vals.get("biggest_reason") or ""),
            biggest_risk=str(vals.get("biggest_risk") or ""),
            resume_angle=str(vals.get("resume_angle") or ""),
            outreach_angle=str(vals.get("outreach_angle") or ""),
            tailor_resume=str(vals.get("tailor_resume") or ""),
            final_recommendation=str(vals.get("final_recommendation") or ""),
        ))
    return out


def save_queue(rows: list[FinderQueueRow], path: Path | None = None) -> Path:
    p = path or QUEUE_PATH
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = _fresh_workbook()
    ws = wb["finder_queue"]
    for i, row in enumerate(rows, start=2):
        d = row.as_dict()
        for col_idx, col in enumerate(COLUMNS, start=1):
            ws.cell(row=i, column=col_idx, value=d.get(col, ""))
        # Color cells
        action = d.get("recommended_next_action") or d.get("recommended_action") or ""
        fill = ACTION_COLORS.get(action)
        if fill:
            ws.cell(row=i, column=COLUMNS.index("recommended_next_action") + 1).fill = fill
            ws.cell(row=i, column=COLUMNS.index("recommended_action") + 1).fill = fill
        cat_fill = CATEGORY_COLORS.get(d.get("role_category") or "")
        if cat_fill:
            ws.cell(row=i, column=COLUMNS.index("role_category") + 1).fill = cat_fill
        int_fill = INTEREST_COLORS.get(d.get("strategic_interest") or "")
        if int_fill:
            ws.cell(row=i, column=COLUMNS.index("strategic_interest") + 1).fill = int_fill
    tmp = p.with_suffix(p.suffix + ".tmp")
    wb.save(str(tmp))
    tmp.replace(p)
    return p


def upsert(rows: list[FinderQueueRow], new_row: FinderQueueRow) -> list[FinderQueueRow]:
    """Idempotent upsert by URL. Preserves all USER_EDITABLE_FIELDS so a
    re-run of the finder doesn't blow away the user's review work."""
    by_url = {r.url: r for r in rows}
    existing = by_url.get(new_row.url)
    if existing:
        for f in USER_EDITABLE_FIELDS:
            existing_val = getattr(existing, f, "")
            if existing_val:
                setattr(new_row, f, existing_val)
    by_url[new_row.url] = new_row
    return list(by_url.values())


# ---------- Diversified ranking ----------

def diversify(rows: list[FinderQueueRow]) -> list[FinderQueueRow]:
    """Round-robin order: best-of-company-A, best-of-company-B, ..., then
    second-best-of-company-A, second-best-of-company-B, etc.

    Side-effects: assigns `company_rank`, `role_rank_within_company`, and
    sets `company_outreach_needed='y'` for any company with 2+ rows."""
    # Group by company_group_id (fall back to company name lowercased)
    groups: dict[str, list[FinderQueueRow]] = {}
    for r in rows:
        gid = r.company_group_id or (r.company or "").strip().lower()
        groups.setdefault(gid, []).append(r)

    # Sort within each company by quick_fit desc (then title alpha for stability)
    for gid, group_rows in groups.items():
        group_rows.sort(key=lambda r: (-int(r.quick_fit or 0), r.title))
        for i, r in enumerate(group_rows, start=1):
            r.role_rank_within_company = i
            r.company_group_id = r.company_group_id or gid
            if len(group_rows) >= 2:
                r.company_outreach_needed = r.company_outreach_needed or "y"

    # Rank companies by their best role's score (tiebreak by name)
    company_best = sorted(
        groups.items(),
        key=lambda kv: (-max(int(r.quick_fit or 0) for r in kv[1]), kv[0]),
    )
    company_rank_map: dict[str, int] = {}
    for rank, (gid, group_rows) in enumerate(company_best, start=1):
        company_rank_map[gid] = rank
        for r in group_rows:
            r.company_rank = rank

    # Round-robin interleave
    ordered: list[FinderQueueRow] = []
    pos = 0
    while True:
        added_this_pass = False
        for gid, _ in company_best:
            group_rows = groups[gid]
            if pos < len(group_rows):
                ordered.append(group_rows[pos])
                added_this_pass = True
        if not added_this_pass:
            break
        pos += 1
    return ordered
