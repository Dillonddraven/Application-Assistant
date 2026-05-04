"""Application tracker — single XLSX, OpenPyXL, source of truth.

Schema:
    job_id | slug | company | title | location | remote_mode | source_url
    fit_score | industry_filter | applied_date | status | last_status_change
    next_followup_date | notes

Status values:
    waiting              — applied, awaiting first response
    interview_pending    — interview scheduled, not yet held
    interview_completed  — interview done, awaiting decision
    rejected             — no further action
    offer                — offer received, awaiting your decision
    accepted             — terminal state, accepted offer
    withdrawn            — you withdrew

Follow-up cadence:
    waiting              -> applied_date  + 7 days
    interview_pending    -> applied_date  + 14 days
    interview_completed  -> last_status_change + 7 days
    others               -> none
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import ROOT

TRACKER_PATH = ROOT / "applications.xlsx"

# v1.1 schema (May 2026): extended for the full apply-track workflow.
# The `submitted_status` column drives the application state machine; the older
# `status` column is preserved for backward read but no longer written.
COLUMNS: list[str] = [
    # Identity
    "application_id", "job_id", "slug",
    # Posting metadata (auto-populated from analyzed.json)
    "company", "role_title", "role_category", "location",
    "source_url", "apply_url",
    "fit_score", "source_confidence",
    # Status / scoring (auto-populated from packet)
    "production_status", "strategic_interest", "recommendation",
    # Local artifacts (auto)
    "packet_folder", "resume_path", "cover_letter_path",
    # Email layer (auto on draft / send)
    "email_created", "email_sent",
    # User-facing application state machine (USER-EDITABLE)
    "submitted_status", "submitted_date",
    "follow_up_date", "outreach_status",
    # Company grouping
    "company_outreach_group",
    # Free-form
    "notes", "next_action",
    # Legacy columns kept so old rows still load gracefully
    "applied_date", "last_status_change", "next_followup_date",
]

# New submitted_status state machine
SUBMITTED_STATUSES: list[str] = [
    "not_started", "packet_ready", "emailed_to_dillon", "submitted",
    "follow_up_needed", "interview", "rejected", "archived", "skipped",
]
NEXT_ACTIONS: list[str] = [
    "review_packet", "submit_portal", "follow_up", "prep_for_interview",
    "wait", "skip",
]
TERMINAL_STATUSES: set[str] = {"rejected", "archived", "skipped"}

# Older status names kept for compatibility — `set_status` will translate them.
LEGACY_STATUSES: list[str] = [
    "waiting", "interview_pending", "interview_completed",
    "rejected", "offer", "accepted", "withdrawn",
]
STATUSES = SUBMITTED_STATUSES + LEGACY_STATUSES

# Cell fills per submitted_status. Pastel = in-progress, muted = terminal.
STATUS_FILLS: dict[str, PatternFill] = {
    "not_started":         PatternFill("solid", fgColor="F2F2F2"),
    "packet_ready":        PatternFill("solid", fgColor="DDE8F7"),
    "emailed_to_dillon":   PatternFill("solid", fgColor="DDEBFB"),
    "submitted":           PatternFill("solid", fgColor="FFF7D6"),
    "follow_up_needed":    PatternFill("solid", fgColor="FFE5B4"),
    "interview":           PatternFill("solid", fgColor="DDF0D2"),
    "rejected":            PatternFill("solid", fgColor="EAD6D6"),
    "archived":            PatternFill("solid", fgColor="E5E5E5"),
    "skipped":             PatternFill("solid", fgColor="E5E5E5"),
    # Legacy values render too
    "waiting":             PatternFill("solid", fgColor="FFF7D6"),
    "interview_pending":   PatternFill("solid", fgColor="DDE8F7"),
    "interview_completed": PatternFill("solid", fgColor="D6EAEA"),
    "offer":               PatternFill("solid", fgColor="DDF0D2"),
    "accepted":            PatternFill("solid", fgColor="9FCF85"),
    "withdrawn":           PatternFill("solid", fgColor="E5E5E5"),
}

HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT = Font(bold=True, color="FFFFFF")


class TrackerError(RuntimeError):
    pass


@dataclass
class Row:
    """One application row, dict-shaped."""
    data: dict[str, Any]

    def get(self, k: str, default: Any = None) -> Any:
        return self.data.get(k, default)

    @property
    def job_id(self) -> str:
        return str(self.data.get("job_id") or "")

    @property
    def slug(self) -> str:
        return str(self.data.get("slug") or "")

    @property
    def status(self) -> str:
        return str(self.data.get("status") or "waiting")


def today_iso() -> str:
    return date.today().isoformat()


def _parse_date(s: Any) -> date | None:
    if not s:
        return None
    if isinstance(s, date):
        return s
    try:
        return datetime.fromisoformat(str(s)).date()
    except ValueError:
        try:
            return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
        except ValueError:
            return None


def followup_for(*, status: str, applied_date: str, last_status_change: str) -> str | None:
    if status in TERMINAL_STATUSES:
        return None
    applied = _parse_date(applied_date)
    last = _parse_date(last_status_change) or applied
    if status == "waiting" and applied:
        return (applied + timedelta(days=7)).isoformat()
    if status == "interview_pending" and applied:
        return (applied + timedelta(days=14)).isoformat()
    if status == "interview_completed" and last:
        return (last + timedelta(days=7)).isoformat()
    return None


def _fresh_workbook() -> Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "applications"
    for i, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"
    widths = {
        "application_id": 14, "job_id": 14, "slug": 38,
        "company": 22, "role_title": 32, "role_category": 16, "location": 18,
        "source_url": 48, "apply_url": 48,
        "fit_score": 9, "source_confidence": 9,
        "production_status": 14, "strategic_interest": 11, "recommendation": 12,
        "packet_folder": 38, "resume_path": 32, "cover_letter_path": 32,
        "email_created": 11, "email_sent": 11,
        "submitted_status": 16, "submitted_date": 12,
        "follow_up_date": 12, "outreach_status": 14,
        "company_outreach_group": 18,
        "notes": 40, "next_action": 16,
        # Legacy
        "applied_date": 12, "last_status_change": 14, "next_followup_date": 14,
    }
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 14)
    return wb


def _load_workbook(path: Path) -> Workbook:
    if not path.exists():
        return _fresh_workbook()
    return openpyxl.load_workbook(str(path))


def _ws(wb: Workbook) -> Worksheet:
    if "applications" in wb.sheetnames:
        return wb["applications"]
    return wb.active


def _row_index_for(ws: Worksheet, *, job_id: str | None = None,
                   slug: str | None = None) -> int | None:
    """Return 1-based row index in the sheet, or None if not found.
    Column positions resolved from COLUMNS so schema changes don't break this."""
    job_id_col = COLUMNS.index("job_id") + 1
    slug_col = COLUMNS.index("slug") + 1
    for r in range(2, ws.max_row + 1):
        ji = ws.cell(row=r, column=job_id_col).value
        sl = ws.cell(row=r, column=slug_col).value
        if job_id and ji == job_id:
            return r
        if slug and sl == slug:
            return r
    return None


def _write_row(ws: Worksheet, row_idx: int, row: dict[str, Any]) -> None:
    for col_idx, col in enumerate(COLUMNS, start=1):
        ws.cell(row=row_idx, column=col_idx, value=row.get(col, ""))
    # Color the submitted_status cell. Fall back to legacy `status` if needed.
    sub_status = str(row.get("submitted_status") or row.get("status") or "")
    fill = STATUS_FILLS.get(sub_status)
    if fill and "submitted_status" in COLUMNS:
        ws.cell(row=row_idx,
                column=COLUMNS.index("submitted_status") + 1).fill = fill


def _row_to_dict(ws: Worksheet, row_idx: int) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for col_idx, col in enumerate(COLUMNS, start=1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if isinstance(val, datetime):
            val = val.date().isoformat()
        elif isinstance(val, date):
            val = val.isoformat()
        out[col] = val if val is not None else ""
    return out


def _save_atomic(wb: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    wb.save(str(tmp))
    tmp.replace(path)


def _new_application_id() -> str:
    """Short stable ID for the application row, e.g. 'a-7f3a2b'."""
    import uuid as _u
    return "a-" + _u.uuid4().hex[:6]


def upsert(*, packet: dict[str, Any], analyzed: dict[str, Any] | None,
           status: str = "packet_ready", notes: str = "",
           path: Path | None = None) -> dict[str, Any]:
    """Insert or update the row for a packet. Idempotent on (job_id, slug).
    Auto-populates posting/packet/path columns; preserves user-edited columns
    (notes, submitted_status, follow_up_date, etc.) on update."""
    # Translate legacy values to the new state machine.
    legacy_to_new = {
        "waiting": "submitted",         # `tracker.upsert(status='waiting')` after approve
        "interview_pending": "interview",
        "interview_completed": "interview",
        "offer": "interview",
        "accepted": "archived",
        "withdrawn": "skipped",
    }
    status = legacy_to_new.get(status, status)
    if status not in STATUSES:
        raise TrackerError(f"unknown status {status!r} (valid: {SUBMITTED_STATUSES})")
    p = path or TRACKER_PATH
    wb = _load_workbook(p)
    ws = _ws(wb)
    job_id = packet.get("job_id") or ""
    slug = packet.get("slug") or ""
    existing_idx = _row_index_for(ws, job_id=job_id, slug=slug)
    today = today_iso()

    # Resolve packet folder + paths if the standard layout is in place.
    from .config import OUTPUTS_DIR
    packet_folder = str(OUTPUTS_DIR / slug) if slug else ""
    employer_pdf = (OUTPUTS_DIR / slug / "employer" / "tailored_resume.pdf") if slug else None
    cover_pdf = (OUTPUTS_DIR / slug / "employer" / "cover_letter.pdf") if slug else None
    resume_path = str(employer_pdf) if employer_pdf and employer_pdf.exists() else ""
    cover_path = str(cover_pdf) if cover_pdf and cover_pdf.exists() else ""

    if existing_idx:
        existing = _row_to_dict(ws, existing_idx)
        application_id = existing.get("application_id") or _new_application_id()
        applied_date = existing.get("applied_date") or today
        # Auto-bump submitted_status from not_started -> packet_ready when a
        # packet is generated, but never overwrite a more advanced state.
        prev = (existing.get("submitted_status") or
                legacy_to_new.get(existing.get("status") or "", "not_started")
                or "not_started")
        progress_order = {s: i for i, s in enumerate(SUBMITTED_STATUSES)}
        if progress_order.get(status, 0) > progress_order.get(prev, 0):
            new_sub_status = status
        else:
            new_sub_status = prev
        last_change = (today if new_sub_status != prev
                       else (existing.get("last_status_change") or today))
        notes_final = notes or existing.get("notes") or ""
        # Preserve user-edited fields
        submitted_date = existing.get("submitted_date") or ""
        follow_up_date = existing.get("follow_up_date") or ""
        outreach_status = existing.get("outreach_status") or ""
        next_action = (existing.get("next_action")
                       or _default_next_action(new_sub_status))
        email_created = existing.get("email_created") or ""
        email_sent = existing.get("email_sent") or ""
        company_group = existing.get("company_outreach_group") or ""
    else:
        application_id = _new_application_id()
        applied_date = today
        last_change = today
        notes_final = notes
        new_sub_status = status
        submitted_date = ""
        follow_up_date = ""
        outreach_status = ""
        next_action = _default_next_action(new_sub_status)
        email_created = ""
        email_sent = ""
        company_group = ""

    a = analyzed or {}
    row = {
        "application_id": application_id,
        "job_id": job_id,
        "slug": slug,
        "company": a.get("company") or "",
        "role_title": a.get("title") or "",
        "role_category": a.get("role_category") or "",
        "location": a.get("location") or "",
        "source_url": a.get("source_url") or "",
        "apply_url": a.get("apply_url") or a.get("source_url") or "",
        "fit_score": a.get("fit_score") or 0,
        "source_confidence": a.get("source_confidence") or "",
        "production_status": packet.get("production_status") or "",
        "strategic_interest": packet.get("strategic_interest") or "",
        "recommendation": packet.get("recommendation") or "",
        "packet_folder": packet_folder,
        "resume_path": resume_path,
        "cover_letter_path": cover_path,
        "email_created": email_created,
        "email_sent": email_sent,
        "submitted_status": new_sub_status,
        "submitted_date": submitted_date,
        "follow_up_date": follow_up_date or (followup_for(
            status="waiting", applied_date=applied_date,
            last_status_change=last_change,
        ) or ""),
        "outreach_status": outreach_status,
        "company_outreach_group": company_group,
        "notes": notes_final,
        "next_action": next_action,
        # Legacy
        "applied_date": applied_date,
        "last_status_change": last_change,
        "next_followup_date": followup_for(
            status="waiting", applied_date=applied_date,
            last_status_change=last_change,
        ) or "",
    }

    target_row = existing_idx or (ws.max_row + 1)
    _write_row(ws, target_row, row)
    _save_atomic(wb, p)
    return row


def _default_next_action(submitted_status: str) -> str:
    return {
        "not_started":      "review_packet",
        "packet_ready":     "review_packet",
        "emailed_to_dillon":"submit_portal",
        "submitted":        "follow_up",
        "follow_up_needed": "follow_up",
        "interview":        "prep_for_interview",
        "rejected":         "skip",
        "archived":         "skip",
        "skipped":          "skip",
    }.get(submitted_status, "wait")


def mark_email_created(slug: str, *, mode: str = "single",
                        path: Path | None = None) -> dict[str, Any] | None:
    """Stamp email_created with today + mode (single|company-bundle).
    Returns updated row or None if no application matches."""
    p = path or TRACKER_PATH
    if not p.exists():
        return None
    wb = _load_workbook(p)
    ws = _ws(wb)
    idx = _row_index_for(ws, slug=slug)
    if idx is None:
        return None
    row = _row_to_dict(ws, idx)
    row["email_created"] = f"{today_iso()} ({mode})"
    if (row.get("submitted_status") in ("not_started", "packet_ready", "")):
        row["submitted_status"] = "emailed_to_dillon"
        row["next_action"] = "submit_portal"
    _write_row(ws, idx, row)
    _save_atomic(wb, p)
    return row


def set_status(*, ident: str, new_status: str, notes: str = "",
               path: Path | None = None) -> dict[str, Any]:
    """Flip submitted_status by job_id or slug. Updates next_action /
    follow_up_date appropriately. Accepts either new (submitted_status) or
    legacy status names; legacy values get translated."""
    legacy_to_new = {
        "waiting": "submitted", "interview_pending": "interview",
        "interview_completed": "interview", "offer": "interview",
        "accepted": "archived", "withdrawn": "skipped",
    }
    new_status = legacy_to_new.get(new_status, new_status)
    if new_status not in STATUSES:
        raise TrackerError(
            f"unknown status {new_status!r} (valid: {SUBMITTED_STATUSES})"
        )
    p = path or TRACKER_PATH
    if not p.exists():
        raise TrackerError(f"no tracker file at {p}; nothing to update.")
    wb = _load_workbook(p)
    ws = _ws(wb)
    idx = _row_index_for(ws, job_id=ident, slug=ident)
    if idx is None:
        # Try matching application_id too
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=COLUMNS.index("application_id") + 1).value == ident:
                idx = r
                break
    if idx is None:
        raise TrackerError(f"no application with job_id, slug, or application_id {ident!r}")
    existing = _row_to_dict(ws, idx)
    today = today_iso()
    prev = existing.get("submitted_status") or ""
    existing["submitted_status"] = new_status
    existing["last_status_change"] = today
    if new_status == "submitted" and not existing.get("submitted_date"):
        existing["submitted_date"] = today
    existing["next_action"] = _default_next_action(new_status)
    existing["next_followup_date"] = followup_for(
        status="waiting" if new_status == "submitted" else "interview_pending"
        if new_status == "interview" else new_status,
        applied_date=existing.get("applied_date") or today,
        last_status_change=today,
    ) or ""
    if notes:
        prefix = (existing.get("notes") or "").rstrip()
        existing["notes"] = (prefix + ("\n" if prefix else "") + f"[{today}] {notes}").strip()
    _write_row(ws, idx, existing)
    _save_atomic(wb, p)
    return existing


def list_rows(*, status: str | None = None, path: Path | None = None) -> list[dict[str, Any]]:
    p = path or TRACKER_PATH
    if not p.exists():
        return []
    wb = _load_workbook(p)
    ws = _ws(wb)
    out: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        d = _row_to_dict(ws, r)
        if not d.get("job_id") and not d.get("slug"):
            continue
        if status:
            row_status = d.get("submitted_status") or d.get("status") or ""
            if row_status != status:
                continue
        out.append(d)
    return out


def due_followups(*, by: date | None = None, path: Path | None = None) -> list[dict[str, Any]]:
    cutoff = by or date.today()
    out: list[dict[str, Any]] = []
    for row in list_rows(path=path):
        if row.get("status") in TERMINAL_STATUSES:
            continue
        d = _parse_date(row.get("next_followup_date"))
        if d and d <= cutoff:
            out.append(row)
    return sorted(out, key=lambda r: r.get("next_followup_date") or "")


def render_table(rows: Iterable[dict[str, Any]]) -> str:
    rows = list(rows)
    if not rows:
        return "(no applications tracked yet)"
    cols = ("ID", "STATUS", "NEXT", "FIT", "COMPANY", "TITLE", "FOLLOWUP")
    lines: list[tuple[str, ...]] = [cols]
    for r in rows:
        sub = r.get("submitted_status") or r.get("status") or ""
        next_a = r.get("next_action") or ""
        title = r.get("role_title") or r.get("title") or ""
        lines.append((
            (r.get("application_id") or "")[:8],
            sub[:18],
            next_a[:18],
            str(r.get("fit_score") or "")[:3],
            (r.get("company") or "")[:22],
            title[:32],
            (r.get("follow_up_date") or r.get("next_followup_date") or "")[:10],
        ))
    widths = [max(len(row[i]) for row in lines) for i in range(len(cols))]
    out: list[str] = []
    for i, row in enumerate(lines):
        out.append("  ".join(cell.ljust(widths[j]) for j, cell in enumerate(row)))
        if i == 0:
            out.append("-" * len(out[-1]))
    return "\n".join(out)
